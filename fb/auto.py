from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import datetime
import time
import pandas as pd
import undetected_chromedriver as uc
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import random
import threading
import os
import csv
import variations
import pyautogui
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import re
from urllib.parse import urlparse
import ssl
from openpyxl import load_workbook

ssl._create_default_https_context = ssl._create_unverified_context
pyautogui.FAILSAFE = False
import sys
import os
import logging

LOG_FILENAME = "console_output.log"

# Shutdown previous handlers and clear the file
logging.shutdown()
if os.path.exists(LOG_FILENAME):
    try:
        os.remove(LOG_FILENAME)
    except Exception as e:
        print(f"Could not clear log file: {e}")

# Re-initialize logger safely
logging.basicConfig(
    filename=LOG_FILENAME,
    filemode='w',
    format='%(asctime)s - %(message)s',
    level=logging.INFO
)

logging.info("🛠 Logger initialized! Log file cleared.")

class Logger:
    def __init__(self, filename="console_output.log"):
        self.terminal = sys.stdout
        self.log = open(filename, "a", encoding="utf-8")
        self.buffer = ""

    def write(self, message):
        self.terminal.write(message)
        self.buffer += message
        while "\n" in self.buffer:
            line, self.buffer = self.buffer.split("\n", 1)

            self.log.write(line + "\n")

    def flush(self):
        if self.buffer:
            self.log.write(self.buffer)
            self.buffer = ""
        self.terminal.flush()
        self.log.flush()

sys.stdout = Logger("console_output.log")
sys.stderr = sys.stdout

print("🛠 Logger initialized! Check console_output.log soon.")



# Load the Excel file
df = pd.read_excel("visited_profiles.xlsx")

# Filter profiles with "pending" status
filtered_df = df[df["Status"] == "pending"]


profiles_visited_counter=0
profile_matched_counter=0
messages_sent_counter=0
already_sent_counter=0
pending_messages_counter=len(filtered_df)
responses_received_counter=0
carry_forward_messages_counter=0


def log_out(driver):
    global profiles_visited_counter
    global profile_matched_counter
    global messages_sent_counter
    global already_sent_counter
    global pending_messages_counter
    global responses_received_counter
    global carry_forward_messages_counter
    profiles_visited_counter=0
    profile_matched_counter=0
    messages_sent_counter=0
    already_sent_counter=0
    pending_messages_counter=len(filtered_df)
    responses_received_counter=0
    carry_forward_messages_counter=0
    try:
        # Wait for the profile button to be visible and clickable
        profile_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//div[@aria-label="Your profile"]'))
        )
        profile_button.click()  # Click the profile button
        time.sleep(2)  # Slight wait after clicking the profile button

    except Exception as e:
        print("Profile button not found or clickable:", e)
    
    try:
        # Wait for the log-out element to be clickable
        log_out_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body//span[contains(text(), 'Log Out')]"))
        )
    except Exception as e:
        print("Log out element not found or clickable:", e)
    print(log_out_element.text)
    log_out_element.click()  # Click the Log Out button
    time.sleep(2)  # Slight wait after logging out


def save_profile_data(url,name, username, page_name):
    # Data to be saved
    data = [url,name, username, page_name, "Yes", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    
    # Specify the CSV file name
    file_name = 'profile_data.csv'
    
    # Check if the file already exists to write header only once
    try:
        with open(file_name, mode='a', newline='') as file:
            writer = csv.writer(file)
            
            # If the file is empty, write the header
            if file.tell() == 0:
                writer.writerow(["Profile_url","Name", "Username", "Page_Name", "Profile_Visited", "Visited_Time"])
            
            # Write the profile data
            writer.writerow(data)
        print(f"Data saved successfully in {file_name}")
    except Exception as e:
        print(f"An error occurred: {e}")

def session_summary(start_time, end_time):
    print("session : ", next_session)
    print("start time : ", start_time)
    print("profiles visited : ", profiles_visited_counter)
    print("profiles matched : ", profile_matched_counter)
    print("messages sent : ", messages_sent_counter)
    print("already sent messages : ", already_sent_counter)
    print("pending messages : ", pending_messages_counter)

# Function to log start and end time in the same row
def log_to_csv(filename, start_time, end_time,username):

    # Load the CSV file
    df = pd.read_csv(filename)

    # Ensure 'session number' column is treated as a number
    df['session number'] = pd.to_numeric(df['session number'], errors='coerce')

    # Find the next session number
    max_session = df['session number'].max()
    global next_session 
    next_session = 1 if pd.isna(max_session) else int(max_session) + 1  # Convert to int

    print(f"Next session number to be entered: {next_session}")

    start_time = datetime.datetime.fromisoformat(start_time)
    end_time = datetime.datetime.fromisoformat(end_time)

    duration = (end_time-start_time).total_seconds()
    # Convert to hours, minutes, and seconds
    hours = int(duration // 3600)
    minutes = int((duration % 3600) // 60)
    seconds = int(duration % 60)

    # Create a new row as a DataFrame
    new_row = pd.DataFrame({'session number': [next_session], 'id_logged' : [username],'Start time': [start_time], 'End time': [end_time], 'profiles_fetched' : [profiles_visited_counter], 'profiles_visited': [profiles_visited_counter], 'profiles_matched':[profile_matched_counter], 'messages_sent':[messages_sent_counter], 'responses_received':[responses_received_counter], 'already_sent':[already_sent_counter], 'carry_forward_profiles':[0], 'carry_forward_messages':[pending_messages_counter], 'duration (sec)': [duration], 'duration (hh:mm:ss)': [f"{hours:02}:{minutes:02}:{seconds:02}"]})  

    # Append the new row using concat
    df = pd.concat([df, new_row], ignore_index=True)

    # Save the updated CSV
    df.to_csv(filename, index=False)

    print(f"Session {next_session} added to {filename}!")

# Define CSV file name
csv_filename = "execution_log.csv"


def human_like_typing(element, text):
    for char in text:
        element.send_keys(char)
        time.sleep(random.uniform(0.1, 0.3))  # Random delay between keystrokes

def can_send_message():
    print("I am in can send message")
    # Load the Excel file
    df = pd.read_excel("visited_profiles.xlsx")

    # Filter rows where Status == "sent"
    filtered_df = df[df["Status"] == "sent"][["Last_Message_Time"]]
 
    if filtered_df.empty:
        return True  # If no previous messages, allow sending

    now = datetime.datetime.now()

    # Count how many messages were sent in the last hour   hours=0.06
    recent_messages = filtered_df[filtered_df["Last_Message_Time"] > now - datetime.timedelta(hours=1)]

    if len(recent_messages) < 2:  # Allow up to 2 messages per hour
        print(f"✅ Allowed: {len(recent_messages)} messages sent in the last hour.")
        return True

    # If 2 messages have already been sent, calculate the wait time
    oldest_time = recent_messages["Last_Message_Time"].min()
    remaining_time = 3600 - (now - oldest_time).total_seconds()
    remaining_minutes = int(remaining_time // 60)
    remaining_seconds = int(remaining_time % 60)

    print(f"⏳ Rate limit reached: Next message in {remaining_minutes} min {remaining_seconds} sec")
    return False


def move_cursor_to_element(element):
    location = element.location
    size = element.size

    # Calculate the center of the element
    x = location['x'] + size['width'] // 2
    y = location['y'] + size['height'] // 2 + 30

    # Move the cursor to the element and click
    pyautogui.moveTo(x, y, duration=0.5)

def open_website(driver):
    websites = [
                    'https://www.w3schools.com/', 
                    'https://egeeksglobal.com/', 
                    'https://newbreak.church/blog/?gad_source=1&gclid=EAIaIQobChMIntaAudXjiwMVxJiDBx1IagDfEAAYASAAEgJ-bPD_BwE',
                    'https://pubmed.ncbi.nlm.nih.gov/', 
                    'https://pubmed.ncbi.nlm.nih.gov/20890834/'
                ]

    # Choose a random website
    random_website = random.choice(websites)
    driver.get(random_website)
    time.sleep(random.randint(3, 10))  # Wait for the new tab to load
    all_windows = driver.window_handles
    main_window = driver.current_window_handle
    # # The new tab should be the last handle in the list (if you haven't opened multiple)
    # new_tab = [handle for handle in all_windows if handle != main_window][0]
    # driver.switch_to.window(new_tab)
    # Perform multiple random scrolls
    num_scrolls = random.randint(2, 5)  # Random number of scrolls
    for _ in range(num_scrolls):
        scroll_amount = random.randint(100, 500)  # Random scroll distance
        driver.execute_script(f"window.scrollBy(0, {scroll_amount});")
        time.sleep(random.uniform(1, 3))  # Random wait time between scrolls

    
    # Check for a search bar
    try:
        search_box = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='search']"))
        )
        search_terms = ["Python", "AI automation", "Selenium tutorial", "Web scraping"]
        search_box.send_keys(random.choice(search_terms))
        search_box.send_keys(Keys.RETURN)
        print("Search bar found and searched something.")
        time.sleep(random.randint(3, 7))

    except:
        print("No search bar found. Clicking a random link.")
        try:
            # Find all links on the page
            links = driver.find_elements(By.TAG_NAME, "a")
            if links:
                random.choice(links).click()
                print("Clicked a random link.")
                time.sleep(random.randint(3, 7))
        except:    
            print("No links found.")
            pass

    # Wait a bit before closing (to simulate user behavior)
    time.sleep(random.randint(3, 5))

def append_to_excel(profiles, messages, url, file_path=r"C:\Users\Mahnoor-Zubair\Desktop\main - Copy\main - Copy\Unread_profiles.xlsx"):
    df = pd.DataFrame({
        "Unread_Profile_Names": profiles,
        "Unread_Profile_Messages": messages,
        "Urls" : url
    })

    if os.path.exists(file_path):
        # Load existing workbook and get last row
        book = load_workbook(file_path)
        sheet = book.active
        start_row = sheet.max_row

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=start_row)
        print("✅ Data appended successfully.")
    else:
        # Create new Excel file
        df.to_excel(file_path, index=False)
        print("✅ File created and data written successfully.")


def unread_profiles(driver):
    try:
        element = driver.find_element(By.CSS_SELECTOR, "div.x9f619.x1ja2u2z")
        if element and "Unread Chats" in element.text.strip():
            print("Correct Element found!")
            match = re.search(r"(\d+) Unread Chats?", element.text.strip())
            if not match:
                print("No unread chats found.")
                return

            unread_count = int(match.group(1))
            print(f"Unread Chats: {unread_count}")
            global responses_received_counter
            responses_received_counter += unread_count

            driver.find_element(By.XPATH, '//div[@role="button" and contains(@aria-label, "Messenger")]').click()
            time.sleep(10)

            unread_profiles_list, unread_messages_list, urls = [], [], []
            profiles = driver.find_elements(By.XPATH, "//div[@role='row' and contains(@class, 'xdj266r') and contains(@class, 'x1n2onr6')]")
            print("Number of profiles:", len(profiles))

            count = 0
            for profile in profiles:
                text = profile.text.strip()
                print(text)
                if "You:" in text:
                    continue

                lines = text.split("\n")
                if len(lines) < 3:
                    continue

                profile_name = lines[0].strip()
                profile_message = lines[1].strip()
                print(f"Profile name: {profile_name}")
                print(f"Profile message: {profile_message}")
                unread_profiles_list.append(profile_name)
                unread_messages_list.append(profile_message)

                try:
                    ActionChains(driver).move_to_element(profile).perform()
                    time.sleep(2)

                    options = profile.find_element(By.XPATH, ".//div[starts-with(@aria-label, 'More options for') and contains(@class, 'x1i10hfl')]")
                    print("found option")
                    print(options.location, options.size)
                    driver.save_screenshot("debug_screen.png")

                    driver.execute_script("arguments[0].scrollIntoView(true);", options)
                    time.sleep(0.5)
                    driver.execute_script("arguments[0].click();", options)
                    print("SVG Element clicked with JS!")

                    time.sleep(3)
                    try:
                        view_profile_button = driver.find_element(By.XPATH, "//div[contains(@class, 'xb57i2i')]//a[.//span[text()='View profile']]")
                        profile_link = view_profile_button.get_attribute("href")
                        if profile_link not in urls:
                            print("Profile link:", profile_link)
                            urls.append(profile_link)
                        else:
                            print("Duplicate profile link ignored.")
                    except Exception as e:
                        print("View Profile button not found.", e)

                except Exception as e:
                    print(f"Error handling profile options: {e}")

                time.sleep(3)
                count += 1
                if count >= unread_count:
                    break
            driver.find_element(By.XPATH, '//div[@role="button" and contains(@aria-label, "Messenger")]').click()
            print("Number of unread profiles (excluding self-sent):", count)
            print("Unread profiles:", unread_profiles_list)
            print("Unread messages:", unread_messages_list)
            print("Profile URLs:", urls)
            append_to_excel(unread_profiles_list, unread_messages_list, urls)
        else:
            print("Element does not contain 'Unread Chats' or wasn't found.")

    except NoSuchElementException as e:
        print("Element not found:", e)

def see_notifications(driver):
    driver.get("https://www.facebook.com/")
    # Find the Notifications button using its class or role attribute
    notifications_button = driver.find_element(By.XPATH, '//div[@role="button" and contains(@aria-label, "Notifications")]')
    move_cursor_to_element(notifications_button)
    # Click the button
    notifications_button.click()
    # Optional: Wait and close browser
    time.sleep(5)
def my_profile(driver):
    driver.get("https://www.facebook.com/")
    # Find the "Your profile" button using XPath
    profile_button = driver.find_element(By.XPATH, '//div[@role="button" and contains(@aria-label, "Your profile")]')
    move_cursor_to_element(profile_button)

    # Click the button
    profile_button.click()

    # Optional: Wait and then close the browser
    time.sleep(5)
def see_menu(driver):
    driver.get("https://www.facebook.com/")
    # Find the "Menu" button using XPath
    menu_button = driver.find_element(By.XPATH, '//div[@role="button" and contains(@aria-label, "Menu")]')
    move_cursor_to_element(menu_button)

    # Click the button
    menu_button.click()

    # Optional: Wait and then close the browser
    time.sleep(5)

def see_memories(driver):
    driver.get("https://www.facebook.com/")
    memories_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//span[text()='Memories']"))
    )
    move_cursor_to_element(memories_button)

    # Click the 'Memories' button
    memories_button.click()
    time.sleep(5)
def see_addsManager(driver):
    driver.get("https://www.facebook.com/")
    addsManager = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//span[text()='Ads Manager']"))
    )
    move_cursor_to_element(addsManager)
    # Click the 'Memories' button
    addsManager.click()
    time.sleep(5)

def see_event(driver):
    driver.get("https://www.facebook.com/")
    event_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//span[text()='Events']"))
    )
    move_cursor_to_element(event_button)

    # Click the 'Memories' button
    event_button.click()
    time.sleep(5)

def see_videos(driver):
    driver.get("https://www.facebook.com/")
    videos_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//span[text()='Video']"))
    )
    move_cursor_to_element(videos_button)

    # Click the 'Memories' button
    videos_button.click()
    time.sleep(5)

def my_variations(driver):
    main_window = driver.current_window_handle
    # Open a new tab
    # Open a new tab
    driver.execute_script("window.open('');")  # Open a blank new tab
    time.sleep(2)  # Wait for the new tab to load
    all_windows = driver.window_handles

    # The new tab should be the last handle in the list (if you haven't opened multiple)
    new_tab = [handle for handle in all_windows if handle != main_window][0]
    driver.switch_to.window(new_tab)
    actions = [open_website,see_menu,see_memories,see_videos, see_addsManager,see_event, see_notifications,my_profile]
    
    random.choice(actions)(driver)
    # Once done, close the new tab
    driver.close()

    # Switch back to the main tab
    driver.switch_to.window(main_window)
    # time.sleep(30)  # Wait for 30 seconds before running again

def scroll(driver):
    # Scroll Downward (Max 5 times)
    down_scrolls = 0
    while down_scrolls < 3:
        driver.execute_script("window.scrollBy(0, 500);")  # Scroll down by 500 pixels
        time.sleep(random.uniform(2, 4))  # Random delay
        down_scrolls += 1

    print(f"Scrolled down {down_scrolls} times.")

    print("Reached the bottom of the page.")

    # Scroll back to the top
    while True:
        driver.execute_script("window.scrollBy(0, -500);")  # Scroll up by 500 pixels
        time.sleep(random.uniform(1, 3))  # Random delay
        
        scroll_position = driver.execute_script("return window.pageYOffset;")
        if scroll_position == 0:  # If we are at the top
            break

    print("Scrolled back to the top.")

def check_photos(driver):
    print("I am checkking photos")
    """Clicks on the photos section."""
    try:
        
        # driver.find_element(By.LINK_TEXT, "Photos").click()
        photos_link = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.LINK_TEXT, "Photos"))
                )
        move_cursor_to_element(photos_link)
        photos_link.click()
        scroll(driver)
        time.sleep(random.randint(3, 7))
    except:
        pass
def check_friends(driver,university_name, univer,username, page_name):
    print("I am checkigf friends")
    """Clicks on the friends section."""
    try:

        # driver.find_element(By.LINK_TEXT, "Friends").click()
        friends_link = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.LINK_TEXT, "Friends"))
                )
        move_cursor_to_element(friends_link)
        friends_link.click()
        scroll(driver)
        time.sleep(random.randint(3, 7))
        driver.execute_script("window.scrollBy(0, 500);")  # Scroll down by 500 pixels
        time.sleep(random.uniform(2, 4))  # Random delay
    
        element = driver.find_elements(By.XPATH, "//div[@class='xq1608w x1p5oq8j']//span[text()='No friends to show']")
        if element:
            print("The element has been found that is no fiends to show")
            return
        else:
            print("The friends are foundd")
            all_links = set()
            last_height = driver.execute_script("return document.body.scrollHeight")
            while True:
                # Get all <a> tags inside the specified <div> using XPath
                links = driver.find_elements(By.XPATH, "//div[@class='x78zum5 x1q0g3np x1a02dak x1qughib']//a")
                
                # Extract and collect href attributes in the set (automatically ensures uniqueness)
                for link in links:
                    href = link.get_attribute("href")
                    if all(x not in href for x in ['/pages/', '/posts/', '/stories/']):
                    # if '/pages/' or '/posts/' or '/stories/' not in href:
                        all_links.add(href)  # Adds only unique links to the set
                
                # Print the links found in this scroll
                print(f"Found {len(links)} links in this scroll.")
                
                # Scroll down by 500 pixels
                driver.execute_script("window.scrollBy(0, 900);")
                time.sleep(random.uniform(2, 4))  # Random delay between scrolls
                
                    # Get the new page height after scrolling
                new_height = driver.execute_script("return document.body.scrollHeight")
                
                # If the height didn't change, we've reached the bottom of the page
                if new_height == last_height:
                    print("Reached the bottom of the page.")
                    break

                # Update the last height to the new height
                last_height = new_height


            # Print all collected unique links
            print(len(all_links))
            print("All collected unique links:")
            for link in all_links:
                print(link)
            for link in all_links:
                print("Visiting the link ", link)
                driver.execute_script("window.open(arguments[0], '_blank');", link)
                time.sleep(10)
                driver.close()
                driver.switch_to.window(driver.window_handles[1])  # Switch to Tab 1
                time.sleep(3)
                about_section = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.LINK_TEXT, "About"))
                )
                pyautogui.moveTo(400, 50, duration=0.5)
                move_cursor_to_element(about_section)
                print("Cursor moved to the element (pyautogui)")
                
                about_section.click()
                target_university = university_name.lower()  # We will match anything containing "Abertay"
                print(f"The target university is : {target_university}")

                time.sleep(random.randint(4,7))
                try:
                    wait = WebDriverWait(driver, 10)
                    print("Compairing the text of the uniersities if the person in university or not")
                    text_ = wait.until(EC.presence_of_element_located(
                    (By.XPATH, "//div[contains(@class, 'x13faqbe')]//span[contains(text(), 'Studies')]")
                    ))#or contains(text(), 'Studied')
                    move_cursor_to_element(text_)
                    print("Cursor moved to the element (pyautogui)")
                    
                    print("✅ Got the text, now comparing with target universities")

                    plain_text = text_.text.lower()
                    print(f"📌 Extracted text: {plain_text}")
                    university_part = plain_text.split("at", 1)[-1].strip()
                    print(f"🔍 Text after 'at': {university_part}")

                    ignore_words = {"university", "college", "at", "in", "of", "the", "studied", "studies", "study", "technology"}
                    words_in_text = set(plain_text.split()) - ignore_words
                    print(f"🔍 Filtered words from extracted text: {words_in_text}")

                    found_match = False
                    print("going into for loop to match university information")
                    for university_name in univer["university"]:
                        print("In side the for loop")
                        words_in_university = set(university_name.lower().split()) - ignore_words
                        matching_words = words_in_text.intersection(words_in_university)
                        print("Now going to check match found or not")
                        if matching_words:
                            print(f"✅ Match found: {university_name} 🎯")

                            global profile_matched_counter
                            profile_matched_counter += 1
                            try:
                        # profile_name_element = driver.find_element(By.XPATH, "//h1[contains(@class, 'html-h1 xdj266r x11i5rnm xat24cr x1mh8g0r xexx8yu x4uap5 x18d9i69 xkhd6sd x1vvkbs x1heor9g x1qlqyl8 x1pd3egz x1a2a7pz')]")
                                profile_name_element = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located(
                                    (By.XPATH, "//div[contains(@class, 'x1e56ztr') or contains(@class, 'x1xmf6yo')]//h1")
                                )
                            )
                                name = profile_name_element.text
                            except:
                                print("Couldnt find the element")
                            find_time = datetime.datetime.now()
                            status = "pending"
                            log_session_visits(link, name, status, find_time)  # Save to Excel
                            global pending_messages_counter
                            pending_messages_counter += 1
                            save_visited_profile(link, name, status, find_time)  # Save to Excel
                            save_profile_data(link ,name, username, page_name)


                            found_match = True
                            break  # Stop checking further universities
                    if not found_match:
                        print(f"⚠ The overview does not show the person is from a target university.")

                except TimeoutException:
                    print("⚠ No university info found in profile. Skipping this person.")


    except:
        print("Frinds not found")
def check_videos(driver):
    print("I am checking videos")
    """Clicks on the videos section."""
    try:
        # driver.find_element(By.LINK_TEXT, "Videos").click()
        Videos_link = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.LINK_TEXT, "Videos"))
                )
        move_cursor_to_element(Videos_link)
        Videos_link.click()
        scroll(driver)
        time.sleep(random.randint(3, 7))
    except:
        pass

def send_request(driver):
    """
    Clicks the 'Message' button and sends a message via a chat input box.
    
    :param driver: Selenium WebDriver instance
    :param message: Message to send
    """
    try:
        print("I am in the try block of the request button")
        request_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@aria-label='Add friend']"))
        )
        move_cursor_to_element(request_button)

        print("I hav ound the request button now i am going to scroll to it")
        # driver.execute_script("arguments[0].scrollIntoView(true);", request_button)
        print("request button found")
        variations.random_cursor_movement()
        time.sleep(3)
        request_button.click()
        time.sleep(5)
        try:
            print("I am entering the etry block of the popup window")
            # Wait for the "OK" button to appear (max wait: 5 seconds)
            ok_button = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//div[@aria-label='OK' and @role='button']"))
            )
            move_cursor_to_element(ok_button)

            print("I have found ok button")
            # Click the button
            ok_button.click()
            print("✅ 'OK' button clicked successfully!")

            # Optional: Wait a moment to let the action complete
            time.sleep(2)

        except Exception as e:
            print("❌ No 'OK' button found or failed to click")
        print("Request button clicked")
        print("Clicked the friend request button.")
    except Exception as e:
        print(f"Error clicking the friend request button: {e}")
        return

def send_message(driver, message):
    """
    Clicks the 'Message' button, checks if the message was already sent, and sends a new one only if necessary.
    
    :param driver: Selenium WebDriver instance
    :param message: Message to send
    """
    global wrote_message
    try:
        print("Trying to locate the 'Message' button...")
        message_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@aria-label='Message']"))
        )
        move_cursor_to_element(message_button)
        print("Cursor moved to the element (pyautogui)")
        
        time.sleep(2)
        print("Message button found, clicking...")
        message_button.click()
    except Exception as e:
        print(f"Error clicking the Message button: {e}")
        return

    # Wait for chat to load
    time.sleep(5)

    try:
        print("Fetching previously sent messages...")
        messages = driver.find_elements(By.XPATH, "//div[contains(@class, 'html-div xdj266r x11i5rnm xat24cr x1mh8g0r xexx8yu x4uap5 x18d9i69 xkhd6sd xeuugli x1vjfegm')]//span")

        # Extract text and clean it up
        sent_messages = [msg.text.strip() for msg in messages if msg.text.strip()]
        print("Retrieved Messages:", sent_messages)  # Debugging

    except Exception as e:
        print(f"Error fetching messages: {e}")
        return

    # Check if the exact message is already sent
    if message.strip() in sent_messages:
        wrote_message=False
        print("Message already sent. Skipping...")
        
        return wrote_message

    try:
        try:
            blocked_element = driver.find_element(By.XPATH, "//div[contains(text(), \"This account can't receive your message\")]")
            print('This id has blocked incoming mesages from a stranger')
            wrote_message = False
            return wrote_message
        except:
            print("Id not blocked the incoming messages from stranger")
        print("Locating the message input box...")
        message_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Message' and @role='textbox']"))
        )
        move_cursor_to_element(message_input)
        print("Cursor moved to the element (pyautogui)")
       
        print("Found message input box, typing message...")
        human_like_typing(message_input, message)
        time.sleep(random.uniform(0.5, 1.5))  # Small delay before hitting Enter
        message_input.send_keys(Keys.RETURN)
        time.sleep(random.uniform(5,6))  # Small delay before hitting Enter
        # actions = ActionChains(driver)
        # actions.move_to_element(message_input).click().send_keys(message.strip() + Keys.RETURN).perform()
        try:
            sent_element = driver.find_element(By.XPATH, "//span[contains(text(), 'Sent')]")
            print("Message sent successfully!")
            wrote_message = True
        except:
            print("The message to this profile couldnt send")
            wrote_message = False
    except Exception as e:
        print(f"Error typing message: {e}")
    return wrote_message

def close_liker_dialog(driver):
    try:
        # Attempt to find and click the close button of the dialog.
        close_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@role='dialog']//div[@aria-label='Close']"))
        )
        move_cursor_to_element(close_button)
        print("Cursor moved to the element (pyautogui)")
        
        close_button.click()
        time.sleep(2)
        print("Dialog closed.")
    except Exception as e:
        # Fallback: Press Escape key to close the dialog.
        driver.find_element(By.TAG_NAME, "body").send_keys("\u001b")
        time.sleep(2)
        print("Dialog closed using Escape key.")


EXCEL_FILE = r"C:\Users\Mahnoor-zubair\Desktop\main - Copy\main - Copy\visited_profiles.xlsx"

from urllib.parse import urlparse, parse_qs, urlencode

def extract_base_url(link):
    parsed = urlparse(link)
    base_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
    query = parse_qs(parsed.query)

    if 'profile.php' in parsed.path and 'id' in query:
        return f"{base_url}?id={query['id'][0]}"
    else:
        return base_url



def load_visited_profiles():
    """Loads visited profiles from the Excel file, ensuring column existence."""
    if os.path.exists(EXCEL_FILE):
        try:
            df = pd.read_excel(EXCEL_FILE)
            if "Links" in df.columns:
                return set(df["Links"].dropna())  # Convert to set for fast lookup
            else:
                print("⚠ Warning: 'Links' column missing in Excel file. Creating a new one.")
                return set()
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return set()
    return set()



def log_session_visits(profile_url, name, status, find_time):
    """Logs the profile visit to the Excel file."""
    if not os.path.exists("log_session_visits.xlsx"):
        df = pd.DataFrame(columns=["Links","Recepient","Find_Time", "Delivered_Status", "Delivered_Time"])  # Create a new file if it doesn't exist
    else:
        df = pd.read_excel("log_session_visits.xlsx")
        # Ensure required columns exist
        for col in ["Links","Recepient","Find_Time", "Delivered_Status", "Delivered_Time"]:
            if col not in df.columns:
                df[col] = []
    df["Delivered_Time"] = pd.to_datetime(df["Delivered_Time"], errors="coerce")
    df["Find_Time"] = pd.to_datetime(df["Find_Time"], errors="coerce")

    # 💥 Convert datetime safely to string or empty
    df["Delivered_Time"] = df["Delivered_Time"].apply(lambda x: x.strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(x) else "")
    df["Find_Time"] = df["Find_Time"].apply(lambda x: x.strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(x) else "")

    delivered_time = datetime.datetime.now()
    
    # Check if the profile already exists
    if profile_url in df["Links"].values:
        df.loc[df["Links"] == profile_url, ["Links","Recepient","Find_Time", "Delivered_Status", "Delivered_Time"]] = [profile_url,name,find_time, status,delivered_time]
    else:
        new_data = pd.DataFrame({"Links" : [profile_url], "Recepient": [name], "Find_Time" : [find_time], "Delivered_Status": [status], "Delivered_Time": [delivered_time]})
        df = pd.concat([df, new_data], ignore_index=True)

    df.to_excel("log_session_visits.xlsx", index=False)


def save_visited_profile(profile_url, name, status, find_time):
    """Appends a newly visited profile to the Excel file, ensuring 'Links' column exists."""
    if not os.path.exists(EXCEL_FILE):
        df = pd.DataFrame(columns=["Links", "Name","Find_Time", "Status", "Last_Message_Time"])  # Create a new file if it doesn't exist
    else:
        df = pd.read_excel(EXCEL_FILE)
        # Ensure required columns exist
        for col in ["Links", "Name","Find_Time", "Status", "Last_Message_Time"]:
            if col not in df.columns:
                df[col] = []
    # Check if the profile already exists
    # Convert to datetime format
    df["Last_Message_Time"] = pd.to_datetime(df["Last_Message_Time"], errors="coerce")
    df["Find_Time"] = pd.to_datetime(df["Find_Time"], errors="coerce")
    # now = datetime.datetime.now()
    # **Only update Last_Message_Time if the message was actually sent**
    last_message_time = datetime.datetime.now()

    if profile_url in df["Links"].values:
        df.loc[df["Links"] == profile_url, ["Name", "Status", "Find_Time","Last_Message_Time"]] = [name, status,find_time, last_message_time]
    else:
        if status == "pending":
            last_message_time = last_message_time
        new_data = pd.DataFrame({"Links": [profile_url], "Name": [name], "Find_Time" : [find_time], "Status": [status], "Last_Message_Time": [last_message_time]})
        df = pd.concat([df, new_data], ignore_index=True)

    df.to_excel(EXCEL_FILE, index=False)


# def profiles_reacted(driver, university_name,univer,username):
#     profile_elements = driver.find_element(By.XPATH,
#         "//div[@class='xb57i2i x1q594ok x5lxg6s x78zum5 xdt5ytf x6ikm8r x1ja2u2z x1pq812k x1rohswg xfk6m8 x1yqm8si xjx87ck xx8ngbg xwo3gff x1n2onr6 x1oyok0e x1odjw0f x1e4zzel x1tbbn4q x1y1aw1k x4uap5 xwib8y2 xkhd6sd']")
#     time.sleep(5)  # Wait for dynamic content to load

#     # Hover over the div
#     ActionChains(driver).move_to_element(profile_elements).perform()
#     # Optional pause to simulate real user
#     time.sleep(2)

#      # Scroll quickly until max height is reached
#     last_height = driver.execute_script("return arguments[0].scrollHeight", profile_elements)

#     while True:
#         driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", profile_elements)
#         time.sleep(1)  # Reduced pause
#         new_height = driver.execute_script("return arguments[0].scrollHeight", profile_elements)

#         if new_height == last_height:
#             break
#         last_height = new_height

#     print("Scrolling complete!")
#     unique_profiles_found = False
#     links = profile_elements.find_elements(By.TAG_NAME, "a")
#     profile_links = [link.get_attribute("href") for link in links if link.get_attribute("href")]
#     print(len(profile_links))
#     # Convert all links to base URLs for comparison
#     profile_base_urls = {extract_base_url(link) for link in profile_links}

#     print("---- Raw Links ----")
#     for link in profile_links[:10]:  # Sample
#         print(link)

#     print("\n---- Extracted Base URLs ----")
#     for base in list(profile_base_urls)[:10]:
#         print(base)
    
#     visited_profiles = load_visited_profiles()
    
#     unique_people_links = profile_base_urls - visited_profiles  # Compare only base URLs

#     if unique_people_links:
#         unique_profiles_found = True 
#         for profile_base_url in unique_people_links:
#             visited_profiles.add(profile_base_url)  # Store only base URLs
            
#             # save_visited_profile(profile_base_url, name, status)  # Save to Excel
#             print(f"✅ Visiting profile: {profile_base_url}")



def profiles_reacted(driver, university_name,univer,username, start_time,page_name):
    status = None
    try:
    #     profile_elements = driver.find_element(By.XPATH,
    #         "//div[@class='xb57i2i x1q594ok x5lxg6s x78zum5 xdt5ytf x6ikm8r x1ja2u2z x1pq812k x1rohswg xfk6m8 x1yqm8si xjx87ck xx8ngbg xwo3gff x1n2onr6 x1oyok0e x1odjw0f x1e4zzel x1tbbn4q x1y1aw1k x4uap5 xwib8y2 xkhd6sd']")
        # Retry mechanism to handle dynamic loading
        attempt = 0
        max_attempts = 5
        profile_elements = None

        while attempt < max_attempts:
            try:
                profile_elements = driver.find_element(By.XPATH, "//div[@class='xb57i2i x1q594ok x5lxg6s x78zum5 xdt5ytf x6ikm8r x1ja2u2z x1pq812k x1rohswg xfk6m8 x1yqm8si xjx87ck xx8ngbg xwo3gff x1n2onr6 x1oyok0e x1odjw0f x1e4zzel x1tbbn4q x1y1aw1k x4uap5 xwib8y2 xkhd6sd']")
                break
            except NoSuchElementException:
                print("🔁 Waiting for scrollable div to appear...")
                time.sleep(3)
                attempt += 1

        time.sleep(5)  # Wait for dynamic content to load

        # Hover over the div
        ActionChains(driver).move_to_element(profile_elements).perform()
        # Optional pause to simulate real user
        time.sleep(2)

        # Scroll quickly until max height is reached
        last_height = driver.execute_script("return arguments[0].scrollHeight", profile_elements)

        while True:
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", profile_elements)
            time.sleep(1)  # Reduced pause
            new_height = driver.execute_script("return arguments[0].scrollHeight", profile_elements)

            if new_height == last_height:
                break
            last_height = new_height

        print("Scrolling complete!")
        unique_profiles_found = False
        links = profile_elements.find_elements(By.TAG_NAME, "a")
        profile_links = [link.get_attribute("href") for link in links if link.get_attribute("href")]
        print(len(profile_links))
        # Convert all links to base URLs for comparison
        
        profile_base_urls = {extract_base_url(link) for link in profile_links}

        print("---- Raw Links ----")
        for link in profile_links[:10]:  # Sample
            print(link)

        print("\n---- Extracted Base URLs ----")
        for base in list(profile_base_urls)[:10]:
            print(base)
        
        visited_profiles = load_visited_profiles()
        
        unique_people_links = profile_base_urls - visited_profiles  # Compare only base URLs

        if unique_people_links:
            unique_profiles_found = True 
            for profile_base_url in unique_people_links:
                # if '/posts/' or '/stories/' or'/pages/' not in profile_base_url:
                if all(x not in profile_base_url for x in ['/pages/', '/posts/', '/stories/']):

                    visited_profiles.add(profile_base_url)  # Store only base URLs
                    
                    # save_visited_profile(profile_base_url, name, status)  # Save to Excel
                    print(f"✅ Visiting profile: {profile_base_url}")

                    main_window = driver.current_window_handle
                    # Open a new tab
                    driver.execute_script("window.open(arguments[0], '_blank');", profile_base_url)
                    global profiles_visited_counter
                    profiles_visited_counter += 1
                    time.sleep(2)  # Wait for the new tab to load
                    all_windows = driver.window_handles

                    # The new tab should be the last handle in the list (if you haven't opened multiple)
                    new_tab = [handle for handle in all_windows if handle != main_window][0]
                    driver.switch_to.window(new_tab)
                    find_time = datetime.datetime.now()
                    try:
                        # profile_name_element = driver.find_element(By.XPATH, "//h1[contains(@class, 'html-h1 xdj266r x11i5rnm xat24cr x1mh8g0r xexx8yu x4uap5 x18d9i69 xkhd6sd x1vvkbs x1heor9g x1qlqyl8 x1pd3egz x1a2a7pz')]")
                        profile_name_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//div[contains(@class, 'x1e56ztr') or contains(@class, 'x1xmf6yo')]//h1")
                )
            )
                        name = profile_name_element.text
                    except:
                        print("Couldnt find the element")
                    # variations.random_cursor_movement()
                    actions = [check_photos, check_videos]
                    random.choice(actions)(driver)
                    driver.execute_script("window.scrollBy(0, 300);")
                    about_section = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.LINK_TEXT, "About"))
                )
                    pyautogui.moveTo(400, 50, duration=0.5)
                    move_cursor_to_element(about_section)
                    print("Cursor moved to the element (pyautogui)")
                    
                    about_section.click()
                    target_university = university_name.lower()  # We will match anything containing "Abertay"
                    print(f"The target university is : {target_university}")

                    time.sleep(random.randint(4,7))

                    try:
                        wait = WebDriverWait(driver, 10)
                        print("Compairing the text of the uniersities if the person in university or not")
                        text_ = wait.until(EC.presence_of_element_located(
                        (By.XPATH, "//div[contains(@class, 'x13faqbe')]//span[contains(text(), 'Studies')]")
                        ))#or contains(text(), 'Studied')
                        move_cursor_to_element(text_)
                        print("Cursor moved to the element (pyautogui)")
                        
                        print("✅ Got the text, now comparing with target universities")

                        plain_text = text_.text.lower()
                        print(f"📌 Extracted text: {plain_text}")
                        university_part = plain_text.split("at", 1)[-1].strip()
                        print(f"🔍 Text after 'at': {university_part}")

                        ignore_words = {"university", "college", "at", "in", "of", "the", "studied", "studies", "study", "technology"}
                        words_in_text = set(plain_text.split()) - ignore_words
                        print(f"🔍 Filtered words from extracted text: {words_in_text}")

                        found_match = False
                        for university_name in univer["university"]:
                            words_in_university = set(university_name.lower().split()) - ignore_words
                            matching_words = words_in_text.intersection(words_in_university)

                            if matching_words:
                                print(f"✅ Match found: {university_name} 🎯")
                                global profile_matched_counter
                                profile_matched_counter += 1
                                # **Check rate limit before sending a message**
                                print("I am before can send message function ")
                                if can_send_message():
                                    unread_profiles(driver)
                                    print("Yes we can send messages and sent mesages variabkle has been initialized")
                                    driver.execute_script("window.scrollTo(0, 0);")
                                    time.sleep(random.randint(2, 5))
                                    send_request(driver)
                                    
                                    wrote_message = send_message(driver, f"Hi {name}, How are you? Are you a student of {university_part}?")
                                    # wrote_message = send_message(driver, "Hello! How are you?")
                                    print("The message was already sent ", wrote_message)
                                    if wrote_message==True:
                                        status = "sent"
                                        save_visited_profile(profile_base_url, name, status, find_time)  # Save to Excel
                                        log_session_visits(profile_base_url, name, status, find_time)  # Save to Excel
                                        global messages_sent_counter
                                        messages_sent_counter += 1
                                    elif wrote_message==False:
                                        status = "already send"
                                        save_visited_profile(profile_base_url, name, status, find_time)  # Save to Excel
                                        log_session_visits(profile_base_url, name, status, find_time)  # Save to Excel
                                        global already_sent_counter
                                        already_sent_counter += 1
                                    # sent_messages.append(datetime.datetime.now())  # Track message timestamp
                                else:
                                    status = "pending"
                                    log_session_visits(profile_base_url, name, status, find_time)  # Save to Excel
                                    global pending_messages_counter
                                    pending_messages_counter += 1
                                    save_visited_profile(profile_base_url, name, status, find_time)  # Save to Excel
                                    
                                found_match = True
                                break  # Stop checking further universities
                                
                        if not found_match:
                            status="not sent"
                            save_visited_profile(profile_base_url, name, status, find_time)  # Save to Excel
                            print(f"⚠ The overview does not show the person is from a target university.")

                    except TimeoutException:
                        status="not sent"
                        save_visited_profile(profile_base_url, name, status, find_time)  # Save to Excel
                        print("⚠ No university info found in profile. Skipping this person.")

                    # except Exception as e:
                    #     print(f"❌ Error during university verification: {str(e)}")

                    save_profile_data(profile_base_url,name, username, page_name)
                    print("The user name is : ",name , "and the Search by is : ",username, "and the page is : ", page_name, "and the status of profile visited is  :  Yes")

                    time.sleep(3)  # Allow about page to load
                    stoper(username,start_time, driver)
                    # Once done, close the new tab
                    # check_friends(driver,university_name, univer,username, page_name)
                    driver.close()

                    # Switch back to the main tab
                    driver.switch_to.window(main_window)
                    time.sleep(6)  # Allow profile to load
        
                    
    except Exception as e:
        print(f"❌ Error extracting or visiting profiles: {e}")
        driver.close()
        # Switch back to the main tab
        driver.switch_to.window(main_window)
        time.sleep(6)  # Allow profile to load
    # Save unmatched profiles to Excel after processing
        # save_profiles_to_excel()
    close_liker_dialog(driver)

def click_all_like_buttons(driver, university_name,visited_profiles,univer, username,start_time,page_name):
    time.sleep(5)
    scroll_pause_time = 2  # Adjust for slower or faster scrolling

    last_height = driver.execute_script("return document.body.scrollHeight")
    count = 0
    while True:
        driver.execute_script("window.scrollBy(0, 500);")  # Scroll down by 300 pixels
        count+=1
        time.sleep(scroll_pause_time)
        
        new_height = driver.execute_script("return window.scrollY + window.innerHeight")
        if new_height == last_height or count == 5:
            break  # Stop when the bottom is reached
        last_height = new_height
    print("I have scrolled to the end")
    stoper(username,start_time, driver)
    try:
        like_buttons = driver.find_elements(
        By.XPATH,
        "//div[@role='button' and .//div[@class='x9f619 x1ja2u2z xzpqnlu x1hyvwdk x14bfe9o xjm9jq1 x6ikm8r x10wlt62 x10l6tqk x1i1rx1s' and normalize-space(text())='All reactions:']]"  # add additional conditions as needed
    )
        
        print(f"Found {len(like_buttons)} like button(s).")
        for index, btn in enumerate(like_buttons):
            send_pending_message(driver)
            try:
                # Scroll into view if necessary
                driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                time.sleep(1)
                if btn.is_displayed():
                    print("The button is displayed")
                    move_cursor_to_element(btn)
                    print("Cursor moved to the element (pyautogui)")
                    time.sleep(1)
                    driver.execute_script("arguments[0].click();", btn)
                    
                    time.sleep(3)
                    print(f"The visited profiles are \n {visited_profiles}")
                    profiles_reacted(driver, university_name,univer,username,start_time,page_name)

                else:
                    print("Button is not displayed")

                print(f"Clicked like button {index+1}")
                time.sleep(2)  # Adjust delay as needed for actions to take effect
            except Exception as e:
                print(f"Error clicking like button {index+1}: {e}")
    except:
        print("No likes button found")    
        pass
def send_pending_message(driver):
    global pending_messages_counter  # ✅ Declare first
    global messages_sent_counter
    global already_sent_counter
# Load the Excel file
    df = pd.read_excel("visited_profiles.xlsx")
    df1 = pd.read_excel("log_session_visits.xlsx")
    # Filter profiles with "pending" status
    filtered_df = df[df["Status"] == "pending"][["Links"]]
    filtered_df1 = df1[df1["Delivered_Status"] == "pending"][["Links"]]
    print("I am filtering only the prending profiles")
    # Store the main window handle
    main_window = driver.current_window_handle 

    # Iterate through pending profiles
    for link in filtered_df["Links"]:
        print("i am in the for loop of the profiles that have been got")
        if can_send_message():
            unread_profiles(driver)
            driver.execute_script("window.open(arguments[0], '_blank');", link)
            time.sleep(2)  # Allow new tab to load

            # Switch to the new tab
            new_tab = [handle for handle in driver.window_handles if handle != main_window][-1]
            driver.switch_to.window(new_tab)
            
            # Send the message
            try:
                profile_name_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, "//div[contains(@class, 'x1e56ztr') or contains(@class, 'x1xmf6yo')]//h1")
            )
        )
                name = profile_name_element.text
            except:
                print("name elemet not found")
            try:
                about_section = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.LINK_TEXT, "About"))
            )
                
                about_section.click()

                try:
                    wait = WebDriverWait(driver, 10)
                    text_ = wait.until(EC.presence_of_element_located(
                    (By.XPATH, "//div[contains(@class, 'x13faqbe')]//span[contains(text(), 'Studies')]")
                    ))#or contains(text(), 'Studied')

                    plain_text = text_.text.lower()
                    print(f"📌 Extracted text: {plain_text}")
                    university_part = plain_text.split(" at ", 1)[-1].strip()
                    print(f"🔍 Text after 'at': {university_part}")
                except:
                    print("No university data found")
            except:
                print("couldnt find the about section")
            wrote_message = send_message(driver, f"Hi {name}, How are you? Are you a student of {university_part}?")
            # wrote_message = send_message(driver, "Hello! How are you?")
            print("The message was already sent ", wrote_message)  
            time.sleep(3)  # Wait for message to send
            if wrote_message==True:
                status = "sent"
                messages_sent_counter += 1
            elif wrote_message==False:
                status = "already send"
                print("The pending nuber of messages are " ,pending_messages_counter) 
                already_sent_counter += 1
            # Update the status and last message time in the DataFrame
            print("Now the staus of profiles will be updated to ", status)
            print(df.head())  # Check first few rows
            df.loc[df["Links"] == link, ["Status", "Last_Message_Time"]] = [status, datetime.datetime.now()]
            df1.loc[df1["Links"] == link, ["Delivered_Status", "Delivered_Time"]] = [status, datetime.datetime.now()]
            print(df.head()) 
            pending_messages_counter -= 1
            print("The status of profile has been updated to ", status)

            # Close the new tab and switch back
            driver.close()
            driver.switch_to.window(main_window)

    # Save the updated DataFrame back to the Excel file
        df.to_excel("visited_profiles.xlsx", index=False)
        df1.to_excel("log_session_visits.xlsx", index=False)
# Function to initialize the Selenium WebDriver
def init_driver():
    options = uc.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--start-maximized")
    # Set the user data directory to store the session
    options.add_argument(r'--user-data-dir=c:\\Users\\Mahnoor-zubair\\Desktop\\sessions')

    # driver = uc.Chrome(version_main=133, options=options)
    driver = uc.Chrome(options=options)
    return driver

from datetime import time as tm
def stoper(username,start_time, driver):
    stop_time = datetime.datetime.combine(datetime.datetime.today().date(), tm(hour=23, minute=00, second=0))

    if datetime.datetime.now() >= stop_time or messages_sent_counter >= 8:
        print("Stop time. Exiting userhit.")
        if os.path.exists("temp_username.txt"):
            os.remove("temp_username.txt")
        # Log session end time
        end_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print("th estart time is ", start_time, "and the end time is ", end_time, "\n ")  
        # Log both times in the same row
        log_to_csv(csv_filename, start_time, end_time, username)
        session_summary(start_time, end_time)
        log_out(driver)    
        raise SystemExit("STOP")
        



# Function to handle the login process
def main_handler(username, password, univer, index, requester):
    # with open("temp_username.txt", "w") as f:
    #     f.write(username)

    status = None  

    driver = init_driver()
    start_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    stoper(username, start_time,driver)
    driver.get("https://www.facebook.com")
    time.sleep(2)  # Allow page to load
    
    try:
        email_input = driver.find_element(By.ID, "email")
        time.sleep(3)
        email_input.send_keys(Keys.CONTROL + "a")  # Select all text
        email_input.send_keys(Keys.BACKSPACE)  # Delete the selected text
        time.sleep(3)
        email_input.send_keys(username)
        time.sleep(3)

        # Locate and enter password
        password_input = driver.find_element(By.ID, "pass")
        time.sleep(3)
        password_input.clear()
        time.sleep(3)
        password_input.send_keys(password)
        time.sleep(3)
        password_input.send_keys(Keys.RETURN)
        time.sleep(5)  # Wait for login to complete
    except:
        pass

    # Check login success
    if "login_attempt" in driver.current_url:
        print(f"Login failed for {username}")
    else:
        print(f"Login successful for {username}")
        time.sleep(5)
        
        try:
            # Wait for the close button to be visible and get its location
            close_button = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Close']"))
            )

            move_cursor_to_element(close_button)
            print("Cursor moved to the close button")

            close_button.click()
            print("Close button clicked successfully!")
            time.sleep(2)
            
        except Exception as e:
            print(f"Error closing post window: {e}")
            time.sleep(2)
        
        time.sleep(5)
        if not univer.empty:
            try:
                for _, uni_row in univer.iterrows():  # ✅ Corrected Loop
                    driver.get("https://www.facebook.com/")
                    university_name = uni_row['university']  # ✅ Now updates correctly for each row
                    print(f"Processing university: {university_name} for {username}")
                    ############################################# THIS IS WHAT I HAVE ADDED ##############################################
                    # time.sleep(555)

                    # my_variations(driver)
                    time.sleep(5)
                    unread_profiles(driver)
                    time.sleep(5)
                    # Search for the university
                    try:
                        search_box = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, "//input[@type='search']"))
                        )
                        move_cursor_to_element(search_box)
                        print("Cursor moved to the close button")

                        driver.execute_script("arguments[0].value = '';", search_box)  # ✅ JavaScript clear
                        search_box.send_keys(Keys.CONTROL + "a")  # ✅ Select all text
                        search_box.send_keys(Keys.BACKSPACE)  # ✅ Delete existing text
                        
                        time.sleep(1)  # Small delay to ensure clearing
                        # university_name = univer.iloc[index]['university']
                        human_like_typing(search_box, university_name)
                        time.sleep(random.uniform(0.5, 1.5))
                        # search_box.send_keys(university_name)
                        search_box.send_keys(Keys.RETURN)
                        time.sleep(random.randint(3, 6))  # Wait for search results
                        print(f"Searching for university: {university_name}")

                    except Exception as e:
                        print(f"Error searching university: {e}")
                    time.sleep(4)
                    try:
                        pages_filter  = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Pages']")))
                        
                        print("Page filter was found")
                        # move_cursor_to_element(pages_filter)
                        print("Cursor moved to the element (pyautogui)")
                        time.sleep(random.randint(2, 5))
                        pages_filter.click()
                        print("Clicked on the pages filter")
                        time.sleep(random.randint(3, 6))
                    except Exception as e:
                        print(f"error clicking on the 'pages' filter : {e}")
                        time.sleep(random.randint(1, 5))

                    try:
                        # variations.random_cursor_movement()
                        # Get all page links
                        page_elements = WebDriverWait(driver, 10).until(
                            EC.presence_of_all_elements_located((By.XPATH, "//a[contains(@href, 'facebook.com')]"))
                        )
                        
                        visited_links = set()
                        matching_links = []


                        for page in page_elements:
                            page_name = page.text.strip().lower()
                            page_link = page.get_attribute('href')
                            # Check if "society" is in the name
                            if ("society" in page_name or 
                                "club" in page_name or 
                                "community" in page_name or 
                                "communities" in page_name):
                                
                                if page_link not in visited_links:
                                    visited_links.add(page_link)
                                    matching_links.append(page_link)
                                    print(f"Society Page Found: {page_name} -> {page_link}")

                                    # driver.get(page_link)
                                    # time.sleep(5)
                        
                        visited_profiles = set()
                        for link in matching_links:
                            try:
                                print(f"\nVisiting page: {link}")
                                driver.get(link)
                                time.sleep(random.randint(2,5))  # Allow time for the page to load
                                

                                # send_pending_message(driver)

                                try:
                                        click_all_like_buttons(driver, university_name,visited_profiles,univer,username,start_time, link)
                                        
                                        # close_liker_dialog(driver)
                                except Exception as e:
                                    print(f"Error finding the post and clicking on the like button {e}")
                                    # Navigate back to the previous page
                                    driver.back()
                                    print("Went back to the previous page")

                                    # Wait for a short period to ensure page is reloaded
                                    time.sleep(2)


                            except Exception as e:
                                print(f"Error visiting {link}: {e}")
                                    
                    except Exception as e:
                        print(f"Error extracting page links: {e}")
                time.sleep(5)
                unread_profiles(driver)
                time.sleep(5)
            except Exception as e:
                print(f"Error fetching university data: {e}")
    
    driver.quit()
    if os.path.exists("temp_username.txt"):
        os.remove("temp_username.txt")
    # Log session end time
    end_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print("th estart time is ", start_time, "and the end time is ", end_time, "\n ")
    
    # Log both times in the same row
    log_to_csv(csv_filename, start_time, end_time, username)
    session_summary(start_time, end_time)
    
if __name__ == "__main__":
    # Load account credentials
    accounts = pd.read_csv("account.csv")
    univer = pd.read_csv("uni.csv")
    
    for i, row in accounts.iterrows():
        
        uname = row['username']
        pas = row['password']
        main_handler(uname, pas, univer, i, "requester")